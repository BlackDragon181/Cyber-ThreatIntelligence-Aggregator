import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import re
from datetime import datetime

class VulnerabilityPrioritizationTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Vulnerability Risk Prioritization Tool")
        self.root.geometry("900x700")
        
        self.input_file = None
        self.output_file = None
        
        self.setup_ui()
        
    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        title_label = ttk.Label(main_frame, text="Vulnerability Risk Prioritization Tool", 
                                font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=10)
        
        # File Selection
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(file_frame, text="Input File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.input_label = ttk.Label(file_frame, text="No file selected", foreground="gray")
        self.input_label.grid(row=0, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.select_input_file).grid(row=0, column=2, padx=5)
        
        # Information Frame - Updated to show automatic intelligence
        info_frame = ttk.LabelFrame(main_frame, text="Intelligent Assessment Criteria", padding="10")
        info_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        info_text = """
The tool automatically analyzes vulnerabilities based on:

✓ CVSS Score > 8.0
✓ Internet Facing detection via Public IP Address
  - Searches: IP, IP_Address, IPv4_Address, asset.display_ipv4_address
  - Automatically identifies public vs private IP ranges
✓ CIA Impact levels (Confidentiality, Integrity, Availability)
✓ Remote Code Execution (RCE) capability
✓ Privilege requirements
✓ User interaction requirements
✓ Public exploit availability
✓ Wormable characteristics
✓ Exploitation likelihood
✓ Indicators of Compromise (IoC)

Each factor = 10 points | Maximum Score = 120 points
        """
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT, foreground="navy").grid(row=0, column=0, sticky=tk.W)
        
        # Process Button
        process_btn = ttk.Button(main_frame, text="Process Vulnerabilities", 
                                command=self.process_vulnerabilities, 
                                style='Accent.TButton')
        process_btn.grid(row=3, column=0, columnspan=3, pady=20)
        
        # Progress
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # Status
        self.status_label = ttk.Label(main_frame, text="Ready", foreground="green")
        self.status_label.grid(row=5, column=0, columnspan=3, pady=5)
        
        # Info Frame
        info_frame = ttk.LabelFrame(main_frame, text="Scoring Information", padding="10")
        info_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        info_text = """
Score Ranges:
• Emergency (110-120): Immediate action required
• Critical (85-110): Address within 24-48 hours
• High (65-85): Address within 1 week
• Medium/Low (Below 65): Schedule for remediation
        """
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT).grid(row=0, column=0, sticky=tk.W)
        
    def select_input_file(self):
        filename = filedialog.askopenfilename(
            title="Select Vulnerability Report",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.input_file = filename
            self.input_label.config(text=filename.split('/')[-1], foreground="black")
            
    def extract_cvss_score(self, row):
        """Extract CVSS score from various column formats"""
        cvss_columns = ['cvss', 'cvss_score', 'cvss_base_score', 'base_score', 'score', 'severity_score']
        
        for col in row.index:
            if any(cvss_col in col.lower() for cvss_col in cvss_columns):
                cvss_str = str(row[col])
                match = re.search(r'\d+\.?\d*', cvss_str)
                if match:
                    return float(match.group())
        return 0.0
    
    def is_public_ip(self, ip_str):
        """Check if an IP address is public (internet-facing)"""
        try:
            import ipaddress
            ip = ipaddress.ip_address(ip_str.strip())
            
            # Check if IP is global (public)
            if ip.is_global:
                return True
            
            # Additional check for public ranges
            if not (ip.is_private or ip.is_loopback or ip.is_link_local or 
                   ip.is_multicast or ip.is_reserved):
                return True
                
        except (ValueError, AttributeError):
            pass
        
        return False
    
    def extract_ip_address(self, row):
        """Extract IP address from various column formats"""
        ip_columns = ['ip', 'ip_address', 'ipv4', 'ipv4_address', 'asset.display_ipv4_address', 
                     'display_ipv4_address', 'host', 'host_ip', 'target', 'target_ip']
        
        for col in row.index:
            col_lower = col.lower()
            # Check if column name contains any IP-related keywords
            if any(ip_col in col_lower for ip_col in ip_columns):
                ip_value = str(row[col]).strip()
                # Basic IP validation pattern
                if re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}
    
    def check_impact_level(self, row, impact_type):
        """Check if impact (Availability/Integrity/Confidentiality) is High"""
        impact_columns = [f'{impact_type}_impact', f'{impact_type}', 'impact']
        
        for col in row.index:
            col_lower = col.lower()
            if any(imp_col in col_lower for imp_col in impact_columns):
                val_str = str(row[col]).lower()
                if 'high' in val_str or 'h' == val_str.strip():
                    return True
        
        # If CVSS >= 7.0, likely high impact
        cvss = self.extract_cvss_score(row)
        if cvss >= 7.0:
            return True
        
        return False
    
    def check_rce(self, row):
        """Determine if vulnerability allows Remote Code Execution"""
        rce_keywords = ['rce', 'remote code execution', 'code execution', 'arbitrary code', 
                       'command execution', 'command injection', 'code injection']
        
        row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
        
        for keyword in rce_keywords:
            if keyword in row_str:
                return True
        
        return False
    
    def check_privileges_required(self, row):
        """Check if privileges are NOT required (returns True if NO privileges needed)"""
        priv_columns = ['privileges', 'privilege_required', 'authentication', 'auth']
        
        for col in row.index:
            col_lower = col.lower()
            if any(priv_col in col_lower for priv_col in priv_columns):
                val_str = str(row[col]).lower()
                # If no privileges needed or authentication not required
                if any(keyword in val_str for keyword in ['none', 'no', 'not required', 'n/a']):
                    return True
                if val_str.strip() in ['n', '0', 'false']:
                    return True
        
        return False
    
    def check_user_interaction(self, row):
        """Check if user interaction is NOT required (returns True if NO interaction needed)"""
        ui_columns = ['user_interaction', 'interaction', 'user_action']
        
        for col in row.index:
            col_lower = col.lower()
            if any(ui_col in col_lower for ui_col in ui_columns):
                val_str = str(row[col]).lower()
                # If no user interaction needed
                if any(keyword in val_str for keyword in ['none', 'no', 'not required', 'n/a']):
                    return True
                if val_str.strip() in ['n', '0', 'false']:
                    return True
        
        return False
    
    def check_exploit_available(self, row):
        """Check if public exploit is available"""
        exploit_keywords = ['exploit', 'poc', 'proof of concept', 'metasploit', 'exploit-db', 
                          'public exploit', 'exploit code']
        exploit_columns = ['exploit', 'exploit_available', 'exploit_code', 'exploitability']
        
        # Check exploit columns
        for col in row.index:
            col_lower = col.lower()
            if any(exp_col in col_lower for exp_col in exploit_columns):
                val_str = str(row[col]).lower()
                if any(keyword in val_str for keyword in ['yes', 'available', 'true', 'public', 'high']):
                    return True
        
        # Check all fields for exploit keywords
        row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
        for keyword in exploit_keywords:
            if keyword in row_str:
                return True
        
        return False
    
    def check_wormable(self, row):
        """Check if vulnerability is wormable"""
        wormable_keywords = ['worm', 'wormable', 'self-propagating', 'propagat']
        
        row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
        
        for keyword in wormable_keywords:
            if keyword in row_str:
                return True
        
        # High CVSS + RCE + No auth + No user interaction often indicates wormable
        if (self.extract_cvss_score(row) >= 9.0 and 
            self.check_rce(row) and 
            self.check_privileges_required(row) and 
            self.check_user_interaction(row)):
            return True
        
        return False
    
    def check_exploitation_likelihood(self, row):
        """Check likelihood of exploitation"""
        likelihood_columns = ['likelihood', 'probability', 'exploitability', 'ease']
        
        for col in row.index:
            col_lower = col.lower()
            if any(like_col in col_lower for like_col in likelihood_columns):
                val_str = str(row[col]).lower()
                if 'high' in val_str or 'likely' in val_str or 'probable' in val_str:
                    return True
        
        # High likelihood if: high CVSS + exploit available + easy to exploit
        if (self.extract_cvss_score(row) >= 8.0 and 
            self.check_exploit_available(row)):
            return True
        
        return False
    
    def check_ioc_detected(self, row):
        """Check for Indicators of Compromise"""
        ioc_keywords = ['ioc', 'indicator', 'compromise', 'detected', 'active', 'exploited', 
                       'in the wild', 'attack detected']
        ioc_columns = ['ioc', 'indicators', 'compromise', 'active_exploitation']
        
        # Check IOC columns
        for col in row.index:
            col_lower = col.lower()
            if any(ioc_col in col_lower for ioc_col in ioc_columns):
                val_str = str(row[col]).lower()
                if any(keyword in val_str for keyword in ['yes', 'true', 'detected', 'active']):
                    return True
        
        # Check all fields for IOC keywords
        row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
        for keyword in ioc_keywords:
            if keyword in row_str:
                return True
        
        return False
    
    def calculate_priority_score(self, row):
        """Calculate priority score based on the scoring matrix with intelligent detection"""
        score = 0
        priority_details = []
        
        # Extract CVSS score
        cvss_score = self.extract_cvss_score(row)
        
        # Priority 1 factors (10 points each)
        # 1. CVSS Score above 8.0
        if cvss_score > 8.0:
            score += 10
            priority_details.append(f"CVSS Score: {cvss_score} > 8.0 [10]")
        
        # 2. Internet Facing/High Exposure (determined by IP address)
        ip_address = self.extract_ip_address(row)
        if self.check_internet_facing(row):
            if ip_address and self.is_public_ip(ip_address):
                score += 10
                priority_details.append(f"Internet Facing - Public IP: {ip_address} [10]")
            else:
                score += 10
                priority_details.append("Internet Facing/High Exposure [10]")
        
        # 3. Availability Impact - High
        if self.check_impact_level(row, 'availability'):
            score += 10
            priority_details.append("High Availability Impact [10]")
        
        # 4. Integrity Impact - High
        if self.check_impact_level(row, 'integrity'):
            score += 10
            priority_details.append("High Integrity Impact [10]")
        
        # 5. Confidentiality Impact - High
        if self.check_impact_level(row, 'confidentiality'):
            score += 10
            priority_details.append("High Confidentiality Impact [10]")
        
        # 6. Is RCE
        if self.check_rce(row):
            score += 10
            priority_details.append("Remote Code Execution (RCE) [10]")
        
        # 7. Privileges NOT Required for exploitation
        if self.check_privileges_required(row):
            score += 10
            priority_details.append("No Privileges Required [10]")
        
        # 8. User Interaction NOT required for exploitation
        if self.check_user_interaction(row):
            score += 10
            priority_details.append("No User Interaction Required [10]")
        
        # 9. Publicly available exploit code
        if self.check_exploit_available(row):
            score += 10
            priority_details.append("Public Exploit Available [10]")
        
        # 10. Is Wormable
        if self.check_wormable(row):
            score += 10
            priority_details.append("Wormable Vulnerability [10]")
        
        # 11. Likelihood for exploitation
        if self.check_exploitation_likelihood(row):
            score += 10
            priority_details.append("High Exploitation Likelihood [10]")
        
        # 12. Indicators of Compromise detected
        if self.check_ioc_detected(row):
            score += 10
            priority_details.append("Indicators of Compromise Detected [10]")
        
        return score, priority_details
    
    def get_risk_rating(self, score):
        """Determine risk rating based on score"""
        if score >= 110:
            return "Emergency"
        elif score >= 85:
            return "Critical"
        elif score >= 65:
            return "High"
        else:
            return "Medium/Low"
    
    def get_priority_level(self, score):
        """Determine priority level"""
        if score >= 110:
            return "Priority 1"
        elif score >= 65:
            return "Priority 2"
        else:
            return "Priority 3"
    
    def process_vulnerabilities(self):
        if not self.input_file:
            messagebox.showerror("Error", "Please select an input file first!")
            return
        
        try:
            self.progress.start()
            self.status_label.config(text="Processing...", foreground="orange")
            self.root.update()
            
            # Read input file
            if self.input_file.endswith('.csv'):
                df = pd.read_csv(self.input_file)
            else:
                df = pd.read_excel(self.input_file)
            
            # Check for required columns
            required_cols = ['CVE', 'IP', 'Asset']
            col_mapping = {}
            
            # Extended IP column search
            ip_col_patterns = ['ip', 'ip_address', 'ipv4', 'ipv4_address', 
                             'asset.display_ipv4_address', 'display_ipv4_address']
            
            for col in required_cols:
                found = False
                for df_col in df.columns:
                    if col.lower() == 'ip':
                        # Special handling for IP columns
                        if any(pattern in df_col.lower() for pattern in ip_col_patterns):
                            col_mapping[col] = df_col
                            found = True
                            break
                    elif col.lower() in df_col.lower():
                        col_mapping[col] = df_col
                        found = True
                        break
                if not found and col != 'IP':  # IP is optional if we can find it in other formats
                    messagebox.showwarning("Warning", 
                        f"Column '{col}' not found. Processing will continue with available data.")
            
            # Calculate scores
            scores = []
            ratings = []
            priorities = []
            details_list = []
            
            for idx, row in df.iterrows():
                score, details = self.calculate_priority_score(row)
                rating = self.get_risk_rating(score)
                priority = self.get_priority_level(score)
                
                scores.append(score)
                ratings.append(rating)
                priorities.append(priority)
                details_list.append("; ".join(details))
            
            # Add new columns
            df['Risk_Score'] = scores
            df['Risk_Rating'] = ratings
            df['Priority_Level'] = priorities
            df['Scoring_Details'] = details_list
            
            # Sort by score (highest first)
            df = df.sort_values('Risk_Score', ascending=False)
            
            # Save output
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"Vulnerability_Prioritization_{timestamp}.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if output_file:
                # Create Excel writer with formatting
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Prioritized Vulnerabilities', index=False)
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Prioritized Vulnerabilities']
                    
                    # Apply conditional formatting colors
                    from openpyxl.styles import PatternFill, Font
                    
                    emergency_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    critical_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    high_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    medium_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    
                    # Find Risk_Rating column
                    rating_col = None
                    for idx, col in enumerate(df.columns, 1):
                        if col == 'Risk_Rating':
                            rating_col = idx
                            break
                    
                    if rating_col:
                        for row_idx in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row_idx, column=rating_col)
                            if cell.value == "Emergency":
                                cell.fill = emergency_fill
                                cell.font = Font(bold=True, color="FFFFFF")
                            elif cell.value == "Critical":
                                cell.fill = critical_fill
                                cell.font = Font(bold=True)
                            elif cell.value == "High":
                                cell.fill = high_fill
                                cell.font = Font(bold=True)
                            elif cell.value == "Medium/Low":
                                cell.fill = medium_fill
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                self.progress.stop()
                self.status_label.config(text=f"Processing complete! Output saved to: {output_file}", 
                                       foreground="green")
                
                # Show summary
                summary = f"""
Processing Complete!

Total Vulnerabilities: {len(df)}
Emergency: {len(df[df['Risk_Rating'] == 'Emergency'])}
Critical: {len(df[df['Risk_Rating'] == 'Critical'])}
High: {len(df[df['Risk_Rating'] == 'High'])}
Medium/Low: {len(df[df['Risk_Rating'] == 'Medium/Low'])}

Output file: {output_file}
                """
                messagebox.showinfo("Success", summary)
            else:
                self.progress.stop()
                self.status_label.config(text="Save cancelled", foreground="orange")
                
        except Exception as e:
            self.progress.stop()
            self.status_label.config(text=f"Error: {str(e)}", foreground="red")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = VulnerabilityPrioritizationTool(root)
    root.mainloop()
, ip_value):
                    return ip_value
        
        return None
    
    def check_internet_facing(self, row):
        """Determine if vulnerability is internet facing based on IP address"""
        # First, try to extract and check IP address
        ip_address = self.extract_ip_address(row)
        
        if ip_address:
            if self.is_public_ip(ip_address):
                return True
        
        # Additional checks for exposure keywords
        exposure_keywords = ['internet', 'external', 'public', 'exposed', 'wan', 'dmz']
        network_keywords = ['network', 'exposure', 'accessibility', 'access', 'zone']
        
        row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
        
        for keyword in exposure_keywords:
            if keyword in row_str:
                return True
        
        # Check if any network-related field contains "high" or "external"
        for col in row.index:
            if any(net_key in col.lower() for net_key in network_keywords):
                val_str = str(row[col]).lower()
                if 'internet' in val_str or 'external' in val_str or 'public' in val_str:
                    return True
        
        return False
    
    def check_impact_level(self, row, impact_type):
        """Check if impact (Availability/Integrity/Confidentiality) is High"""
        impact_columns = [f'{impact_type}_impact', f'{impact_type}', 'impact']
        
        for col in row.index:
            col_lower = col.lower()
            if any(imp_col in col_lower for imp_col in impact_columns):
                val_str = str(row[col]).lower()
                if 'high' in val_str or 'h' == val_str.strip():
                    return True
        
        # If CVSS >= 7.0, likely high impact
        cvss = self.extract_cvss_score(row)
        if cvss >= 7.0:
            return True
        
        return False
    
    def check_rce(self, row):
        """Determine if vulnerability allows Remote Code Execution"""
        rce_keywords = ['rce', 'remote code execution', 'code execution', 'arbitrary code', 
                       'command execution', 'command injection', 'code injection']
        
        row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
        
        for keyword in rce_keywords:
            if keyword in row_str:
                return True
        
        return False
    
    def check_privileges_required(self, row):
        """Check if privileges are NOT required (returns True if NO privileges needed)"""
        priv_columns = ['privileges', 'privilege_required', 'authentication', 'auth']
        
        for col in row.index:
            col_lower = col.lower()
            if any(priv_col in col_lower for priv_col in priv_columns):
                val_str = str(row[col]).lower()
                # If no privileges needed or authentication not required
                if any(keyword in val_str for keyword in ['none', 'no', 'not required', 'n/a']):
                    return True
                if val_str.strip() in ['n', '0', 'false']:
                    return True
        
        return False
    
    def check_user_interaction(self, row):
        """Check if user interaction is NOT required (returns True if NO interaction needed)"""
        ui_columns = ['user_interaction', 'interaction', 'user_action']
        
        for col in row.index:
            col_lower = col.lower()
            if any(ui_col in col_lower for ui_col in ui_columns):
                val_str = str(row[col]).lower()
                # If no user interaction needed
                if any(keyword in val_str for keyword in ['none', 'no', 'not required', 'n/a']):
                    return True
                if val_str.strip() in ['n', '0', 'false']:
                    return True
        
        return False
    
    def check_exploit_available(self, row):
        """Check if public exploit is available"""
        exploit_keywords = ['exploit', 'poc', 'proof of concept', 'metasploit', 'exploit-db', 
                          'public exploit', 'exploit code']
        exploit_columns = ['exploit', 'exploit_available', 'exploit_code', 'exploitability']
        
        # Check exploit columns
        for col in row.index:
            col_lower = col.lower()
            if any(exp_col in col_lower for exp_col in exploit_columns):
                val_str = str(row[col]).lower()
                if any(keyword in val_str for keyword in ['yes', 'available', 'true', 'public', 'high']):
                    return True
        
        # Check all fields for exploit keywords
        row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
        for keyword in exploit_keywords:
            if keyword in row_str:
                return True
        
        return False
    
    def check_wormable(self, row):
        """Check if vulnerability is wormable"""
        wormable_keywords = ['worm', 'wormable', 'self-propagating', 'propagat']
        
        row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
        
        for keyword in wormable_keywords:
            if keyword in row_str:
                return True
        
        # High CVSS + RCE + No auth + No user interaction often indicates wormable
        if (self.extract_cvss_score(row) >= 9.0 and 
            self.check_rce(row) and 
            self.check_privileges_required(row) and 
            self.check_user_interaction(row)):
            return True
        
        return False
    
    def check_exploitation_likelihood(self, row):
        """Check likelihood of exploitation"""
        likelihood_columns = ['likelihood', 'probability', 'exploitability', 'ease']
        
        for col in row.index:
            col_lower = col.lower()
            if any(like_col in col_lower for like_col in likelihood_columns):
                val_str = str(row[col]).lower()
                if 'high' in val_str or 'likely' in val_str or 'probable' in val_str:
                    return True
        
        # High likelihood if: high CVSS + exploit available + easy to exploit
        if (self.extract_cvss_score(row) >= 8.0 and 
            self.check_exploit_available(row)):
            return True
        
        return False
    
    def check_ioc_detected(self, row):
        """Check for Indicators of Compromise"""
        ioc_keywords = ['ioc', 'indicator', 'compromise', 'detected', 'active', 'exploited', 
                       'in the wild', 'attack detected']
        ioc_columns = ['ioc', 'indicators', 'compromise', 'active_exploitation']
        
        # Check IOC columns
        for col in row.index:
            col_lower = col.lower()
            if any(ioc_col in col_lower for ioc_col in ioc_columns):
                val_str = str(row[col]).lower()
                if any(keyword in val_str for keyword in ['yes', 'true', 'detected', 'active']):
                    return True
        
        # Check all fields for IOC keywords
        row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
        for keyword in ioc_keywords:
            if keyword in row_str:
                return True
        
        return False
    
    def calculate_priority_score(self, row):
        """Calculate priority score based on the scoring matrix with intelligent detection"""
        score = 0
        priority_details = []
        
        # Extract CVSS score
        cvss_score = self.extract_cvss_score(row)
        
        # Priority 1 factors (10 points each)
        # 1. CVSS Score above 8.0
        if cvss_score > 8.0:
            score += 10
            priority_details.append(f"CVSS Score: {cvss_score} > 8.0 [10]")
        
        # 2. Internet Facing/High Exposure
        if self.check_internet_facing(row):
            score += 10
            priority_details.append("Internet Facing/High Exposure [10]")
        
        # 3. Availability Impact - High
        if self.check_impact_level(row, 'availability'):
            score += 10
            priority_details.append("High Availability Impact [10]")
        
        # 4. Integrity Impact - High
        if self.check_impact_level(row, 'integrity'):
            score += 10
            priority_details.append("High Integrity Impact [10]")
        
        # 5. Confidentiality Impact - High
        if self.check_impact_level(row, 'confidentiality'):
            score += 10
            priority_details.append("High Confidentiality Impact [10]")
        
        # 6. Is RCE
        if self.check_rce(row):
            score += 10
            priority_details.append("Remote Code Execution (RCE) [10]")
        
        # 7. Privileges NOT Required for exploitation
        if self.check_privileges_required(row):
            score += 10
            priority_details.append("No Privileges Required [10]")
        
        # 8. User Interaction NOT required for exploitation
        if self.check_user_interaction(row):
            score += 10
            priority_details.append("No User Interaction Required [10]")
        
        # 9. Publicly available exploit code
        if self.check_exploit_available(row):
            score += 10
            priority_details.append("Public Exploit Available [10]")
        
        # 10. Is Wormable
        if self.check_wormable(row):
            score += 10
            priority_details.append("Wormable Vulnerability [10]")
        
        # 11. Likelihood for exploitation
        if self.check_exploitation_likelihood(row):
            score += 10
            priority_details.append("High Exploitation Likelihood [10]")
        
        # 12. Indicators of Compromise detected
        if self.check_ioc_detected(row):
            score += 10
            priority_details.append("Indicators of Compromise Detected [10]")
        
        return score, priority_details
    
    def get_risk_rating(self, score):
        """Determine risk rating based on score"""
        if score >= 110:
            return "Emergency"
        elif score >= 85:
            return "Critical"
        elif score >= 65:
            return "High"
        else:
            return "Medium/Low"
    
    def get_priority_level(self, score):
        """Determine priority level"""
        if score >= 110:
            return "Priority 1"
        elif score >= 65:
            return "Priority 2"
        else:
            return "Priority 3"
    
    def process_vulnerabilities(self):
        if not self.input_file:
            messagebox.showerror("Error", "Please select an input file first!")
            return
        
        try:
            self.progress.start()
            self.status_label.config(text="Processing...", foreground="orange")
            self.root.update()
            
            # Read input file
            if self.input_file.endswith('.csv'):
                df = pd.read_csv(self.input_file)
            else:
                df = pd.read_excel(self.input_file)
            
            # Check for required columns
            required_cols = ['CVE', 'IP', 'Asset']
            col_mapping = {}
            
            for col in required_cols:
                found = False
                for df_col in df.columns:
                    if col.lower() in df_col.lower():
                        col_mapping[col] = df_col
                        found = True
                        break
                if not found:
                    messagebox.showwarning("Warning", 
                        f"Column '{col}' not found. Processing will continue with available data.")
            
            # Calculate scores
            scores = []
            ratings = []
            priorities = []
            details_list = []
            
            for idx, row in df.iterrows():
                score, details = self.calculate_priority_score(row)
                rating = self.get_risk_rating(score)
                priority = self.get_priority_level(score)
                
                scores.append(score)
                ratings.append(rating)
                priorities.append(priority)
                details_list.append("; ".join(details))
            
            # Add new columns
            df['Risk_Score'] = scores
            df['Risk_Rating'] = ratings
            df['Priority_Level'] = priorities
            df['Scoring_Details'] = details_list
            
            # Sort by score (highest first)
            df = df.sort_values('Risk_Score', ascending=False)
            
            # Save output
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"Vulnerability_Prioritization_{timestamp}.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if output_file:
                # Create Excel writer with formatting
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Prioritized Vulnerabilities', index=False)
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Prioritized Vulnerabilities']
                    
                    # Apply conditional formatting colors
                    from openpyxl.styles import PatternFill, Font
                    
                    emergency_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    critical_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    high_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    medium_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    
                    # Find Risk_Rating column
                    rating_col = None
                    for idx, col in enumerate(df.columns, 1):
                        if col == 'Risk_Rating':
                            rating_col = idx
                            break
                    
                    if rating_col:
                        for row_idx in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row_idx, column=rating_col)
                            if cell.value == "Emergency":
                                cell.fill = emergency_fill
                                cell.font = Font(bold=True, color="FFFFFF")
                            elif cell.value == "Critical":
                                cell.fill = critical_fill
                                cell.font = Font(bold=True)
                            elif cell.value == "High":
                                cell.fill = high_fill
                                cell.font = Font(bold=True)
                            elif cell.value == "Medium/Low":
                                cell.fill = medium_fill
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                self.progress.stop()
                self.status_label.config(text=f"Processing complete! Output saved to: {output_file}", 
                                       foreground="green")
                
                # Show summary
                summary = f"""
Processing Complete!

Total Vulnerabilities: {len(df)}
Emergency: {len(df[df['Risk_Rating'] == 'Emergency'])}
Critical: {len(df[df['Risk_Rating'] == 'Critical'])}
High: {len(df[df['Risk_Rating'] == 'High'])}
Medium/Low: {len(df[df['Risk_Rating'] == 'Medium/Low'])}

Output file: {output_file}
                """
                messagebox.showinfo("Success", summary)
            else:
                self.progress.stop()
                self.status_label.config(text="Save cancelled", foreground="orange")
                
        except Exception as e:
            self.progress.stop()
            self.status_label.config(text=f"Error: {str(e)}", foreground="red")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = VulnerabilityPrioritizationTool(root)
    root.mainloop()
