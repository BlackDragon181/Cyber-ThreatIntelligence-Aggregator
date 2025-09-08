#!/usr/bin/env python3
"""
Cyber Threat Intelligence (CTI) Aggregator - Clean Working Version
A comprehensive tool for collecting, processing, and analyzing threat intelligence from multiple sources.
"""
# Developer : Kiran Vijay


import asyncio
import json
import hashlib
import re
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, asdict
from urllib.parse import urljoin, urlparse
import xml.etree.ElementTree as ET

# Web framework and HTTP
from flask import Flask, render_template_string, request, jsonify, send_file
import requests
try:
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    RETRY_AVAILABLE = True
except ImportError:
    RETRY_AVAILABLE = False

# Data processing
import pandas as pd
from bs4 import BeautifulSoup
import feedparser

# Export functionality
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    
import tempfile
import os

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class ThreatIntelItem:
    """Data structure for threat intelligence items"""
    id: str
    title: str
    description: str
    category: str  # 'threat_intel', 'advisory', 'ioc'
    source: str
    url: str
    published_date: str
    severity: str = 'unknown'
    tags: List[str] = None
    ioc_type: str = None  # For IOCs: ip, domain, hash, etc.
    ioc_value: str = None
    
    def __post_init__(self):
        if self.tags is None:
            self.tags = []
        # Generate unique ID if not provided
        if not self.id:
            content = f"{self.title}{self.description}{self.source}"
            self.id = hashlib.md5(content.encode()).hexdigest()

class ThreatIntelCollector:
    """Main class for collecting threat intelligence from various sources"""
    
    def __init__(self):
        self.session = self._create_session()
        self.collected_items = []
        self.duplicates_removed = 0
        self.source_stats = {}
        
        # Define comprehensive threat intel sources
        self.sources = {
            # Government & Official Sources
            'cisa_advisories': {
                'url': 'https://www.cisa.gov/cybersecurity-advisories/rss.xml',
                'type': 'rss',
                'category': 'advisory',
                'description': 'CISA Cybersecurity Advisories'
            },
            'us_cert_current': {
                'url': 'https://us-cert.cisa.gov/ncas/current-activity.xml',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'US-CERT Current Activity'
            },
            'nvd_recent': {
                'url': 'https://services.nvd.nist.gov/rest/json/cves/2.0?pubStartDate={}&pubEndDate={}',
                'type': 'nvd_api',
                'category': 'advisory',
                'description': 'NIST NVD Recent CVEs'
            },
            
            # Security News & Research
            'bleeping_computer': {
                'url': 'https://www.bleepingcomputer.com/feed/',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'BleepingComputer Security News'
            },
            'krebs_security': {
                'url': 'https://krebsonsecurity.com/feed/',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'Krebs on Security'
            },
            'dark_reading': {
                'url': 'https://www.darkreading.com/rss.xml',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'Dark Reading'
            },
            'security_week': {
                'url': 'https://www.securityweek.com/feed/',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'SecurityWeek'
            },
            'hacker_news_security': {
                'url': 'https://thehackernews.com/feeds/posts/default',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'The Hacker News'
            },
            
            # Vendor Security Advisories
            'microsoft_security': {
                'url': 'https://www.microsoft.com/security/blog/feed/',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'Microsoft Security Blog'
            },
            'cisco_security': {
                'url': 'https://tools.cisco.com/security/center/psirtrss20/CiscoSecurityAdvisory.xml',
                'type': 'rss',
                'category': 'advisory',
                'description': 'Cisco Security Advisories'
            },
            
            # Security Vendors & Research
            'crowdstrike_blog': {
                'url': 'https://www.crowdstrike.com/blog/feed/',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'CrowdStrike Blog'
            },
            'kaspersky_blog': {
                'url': 'https://securelist.com/feed/',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'Kaspersky Securelist'
            },
            'malwarebytes_blog': {
                'url': 'https://blog.malwarebytes.com/feed/',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'Malwarebytes Labs'
            },
            'checkpoint_research': {
                'url': 'https://research.checkpoint.com/feed/',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'Check Point Research'
            },
            'palo_alto_research': {
                'url': 'https://unit42.paloaltonetworks.com/feed/',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'Palo Alto Unit 42'
            },
            
            # IOC & Malware Feeds
            'threatfox_recent': {
                'url': 'https://threatfox-api.abuse.ch/api/v1/',
                'type': 'threatfox_api',
                'category': 'ioc',
                'description': 'ThreatFox IOCs'
            },
            'malware_bazaar': {
                'url': 'https://mb-api.abuse.ch/api/v1/',
                'type': 'malware_bazaar_api',
                'category': 'ioc',
                'description': 'MalwareBazaar Samples'
            },
            
            # Specialized Threat Intel
            'sans_isc': {
                'url': 'https://isc.sans.edu/rssfeed.xml',
                'type': 'rss',
                'category': 'threat_intel',
                'description': 'SANS Internet Storm Center'
            }
        }
    
    def _create_session(self) -> requests.Session:
        """Create HTTP session with retry strategy"""
        session = requests.Session()
        
        if RETRY_AVAILABLE:
            try:
                # Try newer urllib3 parameter name first
                retry_strategy = Retry(
                    total=3,
                    status_forcelist=[429, 500, 502, 503, 504],
                    allowed_methods=["HEAD", "GET", "OPTIONS", "POST"]
                )
            except TypeError:
                try:
                    # Fallback to older urllib3 parameter name
                    retry_strategy = Retry(
                        total=3,
                        status_forcelist=[429, 500, 502, 503, 504],
                        method_whitelist=["HEAD", "GET", "OPTIONS", "POST"]
                    )
                except TypeError:
                    # Minimal retry strategy if parameters don't work
                    retry_strategy = Retry(total=3)
            
            adapter = HTTPAdapter(max_retries=retry_strategy)
            session.mount("http://", adapter)
            session.mount("https://", adapter)
        
        session.headers.update({
            'User-Agent': 'CTI-Aggregator/1.0 (Security Research)'
        })
        return session
    
    async def collect_all_sources(self) -> List[ThreatIntelItem]:
        """Collect threat intelligence from all configured sources"""
        self.collected_items = []
        self.source_stats = {}
        
        logger.info("Starting comprehensive threat intelligence collection...")
        logger.info(f"Configured sources: {len(self.sources)}")
        
        # Collect from each source
        for source_name, source_config in self.sources.items():
            try:
                logger.info(f"Collecting from {source_config.get('description', source_name)}...")
                start_time = datetime.now()
                
                items = await self._collect_from_source(source_name, source_config)
                
                end_time = datetime.now()
                duration = (end_time - start_time).total_seconds()
                
                self.collected_items.extend(items)
                self.source_stats[source_name] = {
                    'items_collected': len(items),
                    'duration_seconds': duration,
                    'status': 'success' if items or duration < 30 else 'warning',
                    'description': source_config.get('description', source_name)
                }
                
                logger.info(f"✅ {source_config.get('description', source_name)}: {len(items)} items in {duration:.1f}s")
                
            except Exception as e:
                logger.error(f"❌ Error with {source_name}: {str(e)}")
                self.source_stats[source_name] = {
                    'items_collected': 0,
                    'duration_seconds': 0,
                    'status': 'error',
                    'error': str(e),
                    'description': source_config.get('description', source_name)
                }
        
        # Remove duplicates and enhance
        original_count = len(self.collected_items)
        self._remove_duplicates()
        self._categorize_and_enhance()
        
        logger.info(f"Collection complete: {len(self.collected_items)} unique items, {self.duplicates_removed} duplicates removed")
        return self.collected_items
    
    async def _collect_from_source(self, source_name: str, config: Dict) -> List[ThreatIntelItem]:
        """Collect data from a specific source"""
        items = []
        
        try:
            if config['type'] == 'rss':
                items = self._collect_rss(config['url'], config['category'], source_name)
            elif config['type'] == 'nvd_api':
                items = self._collect_nvd(config, source_name)
            elif config['type'] == 'threatfox_api':
                items = self._collect_threatfox(config, source_name)
            elif config['type'] == 'malware_bazaar_api':
                items = self._collect_malware_bazaar(config, source_name)
            else:
                logger.warning(f"Unknown source type: {config['type']} for {source_name}")
        except Exception as e:
            logger.error(f"Error collecting from {source_name}: {str(e)}")
            items = []
        
        return items
    
    def _collect_rss(self, url: str, category: str, source: str) -> List[ThreatIntelItem]:
        """Collect from RSS feeds"""
        items = []
        try:
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            
            feed = feedparser.parse(response.content)
            
            for entry in feed.entries[:30]:  # Limit to recent 30 items
                title = entry.get('title', 'No Title')
                
                # Try multiple description fields
                description = (
                    entry.get('summary', '') or 
                    entry.get('description', '') or 
                    entry.get('content', [{}])[0].get('value', '') if entry.get('content') else ''
                )
                
                # Clean HTML from description
                description = self._clean_html(description)
                
                # Extract publication date
                pub_date = entry.get('published') or entry.get('updated') or datetime.now().isoformat()
                
                # Enhanced severity detection
                severity = self._extract_severity(title + ' ' + description)
                
                # Enhanced tags
                tags = self._extract_tags(title + ' ' + description)
                
                item = ThreatIntelItem(
                    id='',
                    title=title,
                    description=description[:500] + '...' if len(description) > 500 else description,
                    category=category,
                    source=source,
                    url=entry.get('link', ''),
                    published_date=pub_date,
                    severity=severity,
                    tags=tags
                )
                items.append(item)
                
        except Exception as e:
            logger.error(f"Error collecting RSS from {url}: {str(e)}")
        
        return items
    
    def _collect_nvd(self, config: Dict, source: str) -> List[ThreatIntelItem]:
        """Collect from NVD (National Vulnerability Database)"""
        items = []
        try:
            # Get CVEs from last 7 days
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            
            url = config['url'].format(
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d')
            )
            
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            
            data = response.json()
            
            for vuln in data.get('vulnerabilities', [])[:50]:  # Limit to 50 items
                cve = vuln.get('cve', {})
                descriptions = cve.get('descriptions', [])
                description = descriptions[0].get('value', '') if descriptions else ''
                
                metrics = cve.get('metrics', {})
                severity = 'unknown'
                if 'cvssMetricV31' in metrics and metrics['cvssMetricV31']:
                    severity = metrics['cvssMetricV31'][0]['cvssData'].get('baseSeverity', 'unknown')
                
                item = ThreatIntelItem(
                    id='',
                    title=cve.get('id', 'Unknown CVE'),
                    description=description,
                    category='advisory',
                    source=source,
                    url=f"https://nvd.nist.gov/vuln/detail/{cve.get('id', '')}",
                    published_date=cve.get('published', datetime.now().isoformat()),
                    severity=severity.lower() if severity else 'unknown',
                    tags=self._extract_cve_tags(description)
                )
                items.append(item)
                
        except Exception as e:
            logger.error(f"Error collecting from NVD: {str(e)}")
        
        return items
    
    def _collect_threatfox(self, config: Dict, source: str) -> List[ThreatIntelItem]:
        """Collect IOCs from ThreatFox"""
        items = []
        try:
            payload = {
                'query': 'get_iocs',
                'days': 7
            }
            
            response = self.session.post(
                config['url'],
                json=payload,
                headers={'Content-Type': 'application/json'},
                timeout=30
            )
            response.raise_for_status()
            
            data = response.json()
            
            if data.get('query_status') == 'ok':
                for ioc_data in data.get('data', [])[:100]:  # Limit to 100 IOCs
                    item = ThreatIntelItem(
                        id='',
                        title=f"IOC: {ioc_data.get('ioc', 'Unknown')}",
                        description=f"Threat: {ioc_data.get('threat_type', 'Unknown')} | Tags: {', '.join(ioc_data.get('tags', []))}",
                        category='ioc',
                        source=source,
                        url=f"https://threatfox.abuse.ch/ioc/{ioc_data.get('id', '')}",
                        published_date=ioc_data.get('first_seen', datetime.now().isoformat()),
                        severity=self._malware_to_severity(ioc_data.get('malware', '')),
                        tags=ioc_data.get('tags', []),
                        ioc_type=ioc_data.get('ioc_type', 'unknown'),
                        ioc_value=ioc_data.get('ioc', '')
                    )
                    items.append(item)
                    
        except Exception as e:
            logger.error(f"Error collecting from ThreatFox: {str(e)}")
        
        return items
    
    def _collect_malware_bazaar(self, config: Dict, source: str) -> List[ThreatIntelItem]:
        """Collect malware samples from MalwareBazaar"""
        items = []
        try:
            payload = {
                'query': 'get_recent',
                'selector': '50'
            }
            
            response = self.session.post(
                config['url'],
                json=payload,
                headers={'Content-Type': 'application/json'},
                timeout=30
            )
            response.raise_for_status()
            
            data = response.json()
            
            if data.get('query_status') == 'ok':
                for sample in data.get('data', [])[:30]:  # Limit to 30 samples
                    item = ThreatIntelItem(
                        id='',
                        title=f"Malware Sample: {sample.get('file_name', 'Unknown')}",
                        description=f"Type: {sample.get('file_type', 'Unknown')} | Size: {sample.get('file_size', 'Unknown')} bytes",
                        category='ioc',
                        source=source,
                        url=f"https://bazaar.abuse.ch/sample/{sample.get('sha256_hash', '')}",
                        published_date=sample.get('first_seen', datetime.now().isoformat()),
                        severity='high',
                        tags=sample.get('tags', []),
                        ioc_type='hash',
                        ioc_value=sample.get('sha256_hash', '')
                    )
                    items.append(item)
                    
        except Exception as e:
            logger.error(f"Error collecting from MalwareBazaar: {str(e)}")
        
        return items
    
    def _remove_duplicates(self):
        """Remove duplicate items based on content similarity"""
        seen_hashes = set()
        unique_items = []
        
        for item in self.collected_items:
            content_hash = hashlib.md5(
                f"{item.title}{item.description}".encode()
            ).hexdigest()
            
            if content_hash not in seen_hashes:
                seen_hashes.add(content_hash)
                unique_items.append(item)
            else:
                self.duplicates_removed += 1
        
        self.collected_items = unique_items
    
    def _categorize_and_enhance(self):
        """Enhance items with better categorization and metadata"""
        for item in self.collected_items:
            # Enhance IOC detection
            if item.category == 'threat_intel':
                if self._contains_ioc_patterns(item.description):
                    item.category = 'ioc'
            
            # Enhance severity detection
            if item.severity == 'unknown':
                item.severity = self._extract_severity(item.title + ' ' + item.description)
            
            # Enhance tags
            additional_tags = self._extract_tags(item.title + ' ' + item.description)
            item.tags.extend([tag for tag in additional_tags if tag not in item.tags])
    
    def _clean_html(self, text: str) -> str:
        """Remove HTML tags and clean text"""
        if not text:
            return ""
        soup = BeautifulSoup(text, 'html.parser')
        return soup.get_text().strip()
    
    def _extract_severity(self, text: str) -> str:
        """Extract severity from text content"""
        text_lower = text.lower()
        if any(word in text_lower for word in ['critical', 'severe', 'emergency']):
            return 'critical'
        elif any(word in text_lower for word in ['high', 'important', 'major']):
            return 'high'
        elif any(word in text_lower for word in ['medium', 'moderate', 'warning']):
            return 'medium'
        elif any(word in text_lower for word in ['low', 'minor', 'info']):
            return 'low'
        return 'unknown'
    
    def _extract_tags(self, text: str) -> List[str]:
        """Extract relevant tags from text"""
        tags = []
        text_lower = text.lower()
        
        # Common threat categories
        threat_keywords = {
            'malware': ['malware', 'trojan', 'virus', 'backdoor', 'rootkit'],
            'ransomware': ['ransomware', 'encrypt', 'ransom'],
            'phishing': ['phishing', 'spear phishing', 'credential harvest'],
            'apt': ['apt', 'advanced persistent threat', 'nation state'],
            'vulnerability': ['vulnerability', 'cve-', 'exploit', 'zero-day'],
            'ddos': ['ddos', 'denial of service', 'botnet'],
            'data_breach': ['data breach', 'data leak', 'information disclosure']
        }
        
        for tag, keywords in threat_keywords.items():
            if any(keyword in text_lower for keyword in keywords):
                tags.append(tag)
        
        return tags
    
    def _extract_cve_tags(self, description: str) -> List[str]:
        """Extract tags specific to CVE descriptions"""
        tags = []
        desc_lower = description.lower()
        
        if 'remote code execution' in desc_lower or 'rce' in desc_lower:
            tags.append('rce')
        if 'sql injection' in desc_lower:
            tags.append('sqli')
        if 'cross-site scripting' in desc_lower or 'xss' in desc_lower:
            tags.append('xss')
        if 'buffer overflow' in desc_lower:
            tags.append('buffer_overflow')
        if 'privilege escalation' in desc_lower:
            tags.append('privilege_escalation')
        
        return tags
    
    def _contains_ioc_patterns(self, text: str) -> bool:
        """Check if text contains IOC patterns"""
        ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
        domain_pattern = r'\b[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b'
        hash_patterns = [
            r'\b[a-fA-F0-9]{32}\b',  # MD5
            r'\b[a-fA-F0-9]{40}\b',  # SHA1
            r'\b[a-fA-F0-9]{64}\b'   # SHA256
        ]
        
        patterns = [ip_pattern, domain_pattern] + hash_patterns
        return any(re.search(pattern, text) for pattern in patterns)
    
    def _malware_to_severity(self, malware_type: str) -> str:
        """Map malware type to severity"""
        if not malware_type:
            return 'medium'
        
        malware_lower = malware_type.lower()
        if any(word in malware_lower for word in ['ransomware', 'rootkit', 'backdoor']):
            return 'critical'
        elif any(word in malware_lower for word in ['trojan', 'stealer', 'keylogger']):
            return 'high'
        else:
            return 'medium'

class DataExporter:
    """Handle data export functionality"""
    
    @staticmethod
    def to_excel(items: List[ThreatIntelItem], filename: str = None) -> str:
        """Export data to Excel format"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is not available. Please install it with: pip install openpyxl")
            
        if filename is None:
            filename = f"threat_intel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Convert to DataFrame
        data = []
        for item in items:
            row = asdict(item)
            row['tags'] = ', '.join(row['tags']) if row['tags'] else ''
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Create temporary file
        temp_file = os.path.join(tempfile.gettempdir(), filename)
        
        # Write to Excel with formatting
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Threat Intelligence', index=False)
            
            # Format the worksheet
            worksheet = writer.sheets['Threat Intelligence']
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return temp_file
    
    @staticmethod
    def to_json(items: List[ThreatIntelItem], filename: str = None) -> str:
        """Export data to JSON format"""
        if filename is None:
            filename = f"threat_intel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        # Convert to JSON-serializable format
        data = {
            'export_timestamp': datetime.now().isoformat(),
            'total_items': len(items),
            'categories': {},
            'items': [asdict(item) for item in items]
        }
        
        # Add category statistics
        for item in items:
            category = item.category
            if category not in data['categories']:
                data['categories'][category] = 0
            data['categories'][category] += 1
        
        temp_file = os.path.join(tempfile.gettempdir(), filename)
        
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        
        return temp_file

# Flask Web Application
app = Flask(__name__)
app.secret_key = 'cti-aggregator-secret-key-2024'

# Global collector instance
collector = ThreatIntelCollector()

@app.route('/')
def index():
    """Main dashboard page"""
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/sources')
def get_sources():
    """Get information about available sources"""
    try:
        sources_info = []
        for source_name, config in collector.sources.items():
            sources_info.append({
                'name': source_name,
                'description': config.get('description', source_name),
                'type': config['type'],
                'category': config['category'],
                'url': config['url'] if 'api' not in config['type'] else 'API Endpoint'
            })
        
        return jsonify({
            'success': True,
            'total_sources': len(sources_info),
            'sources': sources_info
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/collect', methods=['POST'])
def collect_intelligence():
    """API endpoint to collect threat intelligence"""
    try:
        # Run collection in async context
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        items = loop.run_until_complete(collector.collect_all_sources())
        loop.close()
        
        # Prepare response data
        response_data = {
            'success': True,
            'total_items': len(items),
            'duplicates_removed': collector.duplicates_removed,
            'categories': {},
            'sources': {},
            'severities': {},
            'source_stats': collector.source_stats,
            'items': []
        }
        
        # Process items for response
        for item in items:
            # Count by category
            if item.category not in response_data['categories']:
                response_data['categories'][item.category] = 0
            response_data['categories'][item.category] += 1
            
            # Count by source
            if item.source not in response_data['sources']:
                response_data['sources'][item.source] = 0
            response_data['sources'][item.source] += 1
            
            # Count by severity
            if item.severity not in response_data['severities']:
                response_data['severities'][item.severity] = 0
            response_data['severities'][item.severity] += 1
            
            # Add item to response
            item_dict = asdict(item)
            response_data['items'].append(item_dict)
        
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Error during collection: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/export/<format>')
def export_data(format):
    """Export collected data in specified format"""
    try:
        if not collector.collected_items:
            return jsonify({'error': 'No data to export. Please collect data first.'}), 400
        
        if format.lower() == 'excel':
            if not EXCEL_AVAILABLE:
                return jsonify({'error': 'Excel export not available. Please install openpyxl: pip install openpyxl'}), 400
            file_path = DataExporter.to_excel(collector.collected_items)
            return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))
        
        elif format.lower() == 'json':
            file_path = DataExporter.to_json(collector.collected_items)
            return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))
        
        else:
            return jsonify({'error': 'Unsupported format. Use "excel" or "json".'}), 400
            
    except Exception as e:
        logger.error(f"Error during export: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/filter')
def filter_data():
    """Filter collected data based on parameters"""
    try:
        category = request.args.get('category')
        source = request.args.get('source')
        severity = request.args.get('severity')
        search = request.args.get('search', '').lower()
        
        filtered_items = collector.collected_items
        
        if category:
            filtered_items = [item for item in filtered_items if item.category == category]
        
        if source:
            filtered_items = [item for item in filtered_items if item.source == source]
        
        if severity:
            filtered_items = [item for item in filtered_items if item.severity == severity]
        
        if search:
            filtered_items = [
                item for item in filtered_items
                if search in item.title.lower() or search in item.description.lower()
            ]
        
        return jsonify({
            'success': True,
            'total_filtered': len(filtered_items),
            'items': [asdict(item) for item in filtered_items]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# HTML Template for the web interface
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cyber Threat Intelligence Aggregator</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
            padding: 20px;
        }
        
        .header {
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(10px);
            border-radius: 15px;
            padding: 30px;
            margin-bottom: 30px;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
            text-align: center;
        }
        
        .header h1 {
            color: #2c3e50;
            font-size: 2.5rem;
            margin-bottom: 10px;
        }
        
        .header p {
            color: #7f8c8d;
            font-size: 1.1rem;
            margin-bottom: 15px;
        }
        
        .source-tags {
            display: flex;
            gap: 10px;
            justify-content: center;
            flex-wrap: wrap;
            font-size: 0.9rem;
        }
        
        .source-tag {
            background: rgba(52, 152, 219, 0.1);
            padding: 5px 12px;
            border-radius: 15px;
            border: 1px solid rgba(52, 152, 219, 0.3);
        }
        
        .controls {
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(10px);
            border-radius: 15px;
            padding: 25px;
            margin-bottom: 30px;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        }
        
        .button-group {
            display: flex;
            gap: 15px;
            flex-wrap: wrap;
            align-items: center;
            justify-content: center;
            margin-bottom: 20px;
        }
        
        .btn {
            padding: 12px 24px;
            border: none;
            border-radius: 25px;
            cursor: pointer;
            font-size: 1rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            transition: all 0.3s ease;
        }
        
        .btn-primary {
            background: linear-gradient(45deg, #3498db, #2980b9);
            color: white;
            box-shadow: 0 4px 15px rgba(52, 152, 219, 0.3);
        }
        
        .btn-primary:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(52, 152, 219, 0.4);
        }
        
        .btn-success {
            background: linear-gradient(45deg, #27ae60, #229954);
            color: white;
            box-shadow: 0 4px 15px rgba(39, 174, 96, 0.3);
        }
        
        .btn-success:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(39, 174, 96, 0.4);
        }
        
        .btn-warning {
            background: linear-gradient(45deg, #f39c12, #e67e22);
            color: white;
            box-shadow: 0 4px 15px rgba(243, 156, 18, 0.3);
        }
        
        .btn-warning:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(243, 156, 18, 0.4);
        }
        
        .btn-info {
            background: linear-gradient(45deg, #8e44ad, #9b59b6);
            color: white;
            box-shadow: 0 4px 15px rgba(142, 68, 173, 0.3);
        }
        
        .btn-info:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(142, 68, 173, 0.4);
        }
        
        .btn:disabled {
            background: #bdc3c7;
            cursor: not-allowed;
            transform: none;
            box-shadow: none;
        }
        
        .sources-view {
            margin-bottom: 20px;
            background: rgba(255,255,255,0.9);
            padding: 20px;
            border-radius: 10px;
            display: none;
        }
        
        .sources-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 15px;
        }
        
        .source-item {
            background: white;
            padding: 15px;
            border-radius: 8px;
            border-left: 4px solid #3498db;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }
        
        .source-name {
            font-weight: 600;
            color: #2c3e50;
            margin-bottom: 5px;
        }
        
        .source-meta {
            font-size: 0.8rem;
            color: #7f8c8d;
        }
        
        .source-meta span {
            background: #ecf0f1;
            padding: 2px 6px;
            border-radius: 10px;
            margin-right: 5px;
        }
        
        .filters {
            display: flex;
            gap: 15px;
            flex-wrap: wrap;
            align-items: center;
        }
        
        .filter-group {
            display: flex;
            flex-direction: column;
            gap: 5px;
        }
        
        .filter-group label {
            font-weight: 600;
            color: #2c3e50;
            font-size: 0.9rem;
        }
        
        .filter-group select,
        .filter-group input {
            padding: 8px 12px;
            border: 2px solid #ecf0f1;
            border-radius: 8px;
            font-size: 0.9rem;
            transition: border-color 0.3s ease;
        }
        
        .filter-group select:focus,
        .filter-group input:focus {
            outline: none;
            border-color: #3498db;
        }
        
        .stats {
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(10px);
            border-radius: 15px;
            padding: 25px;
            margin-bottom: 30px;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        }
        
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 15px;
        }
        
        .stat-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 12px;
            text-align: center;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1);
            transition: transform 0.3s ease;
        }
        
        .stat-card:hover {
            transform: translateY(-3px);
        }
        
        .stat-card h3 {
            font-size: 1.8rem;
            margin-bottom: 5px;
        }
        
        .stat-card p {
            font-size: 0.9rem;
            opacity: 0.9;
        }
        
        .results {
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(10px);
            border-radius: 15px;
            padding: 25px;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        }
        
        .results h2 {
            color: #2c3e50;
            margin-bottom: 20px;
            font-size: 1.8rem;
        }
        
        .loading {
            text-align: center;
            padding: 40px;
            color: #7f8c8d;
            font-size: 1.1rem;
        }
        
        .spinner {
            border: 4px solid #ecf0f1;
            border-left: 4px solid #3498db;
            border-radius: 50%;
            width: 50px;
            height: 50px;
            animation: spin 1s linear infinite;
            margin: 0 auto 20px;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        .intel-grid {
            display: grid;
            gap: 20px;
            grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
        }
        
        .intel-item {
            background: white;
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1);
            border-left: 5px solid #3498db;
            transition: transform 0.3s ease, box-shadow 0.3s ease;
        }
        
        .intel-item:hover {
            transform: translateY(-3px);
            box-shadow: 0 8px 25px rgba(0, 0, 0, 0.15);
        }
        
        .intel-item.category-threat_intel {
            border-left-color: #e74c3c;
        }
        
        .intel-item.category-advisory {
            border-left-color: #f39c12;
        }
        
        .intel-item.category-ioc {
            border-left-color: #9b59b6;
        }
        
        .intel-header {
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            margin-bottom: 15px;
            gap: 15px;
        }
        
        .intel-title {
            font-size: 1.1rem;
            font-weight: 600;
            color: #2c3e50;
            flex: 1;
        }
        
        .intel-meta {
            display: flex;
            flex-direction: column;
            align-items: flex-end;
            gap: 5px;
        }
        
        .category-badge,
        .severity-badge,
        .source-badge {
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 0.8rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        
        .category-badge {
            background: #ecf0f1;
            color: #2c3e50;
        }
        
        .severity-badge.critical {
            background: #e74c3c;
            color: white;
        }
        
        .severity-badge.high {
            background: #f39c12;
            color: white;
        }
        
        .severity-badge.medium {
            background: #f1c40f;
            color: #2c3e50;
        }
        
        .severity-badge.low {
            background: #27ae60;
            color: white;
        }
        
        .severity-badge.unknown {
            background: #95a5a6;
            color: white;
        }
        
        .source-badge {
            background: #3498db;
            color: white;
            font-size: 0.7rem;
        }
        
        .intel-description {
            color: #555;
            margin-bottom: 15px;
            line-height: 1.6;
        }
        
        .intel-tags {
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-bottom: 15px;
        }
        
        .tag {
            background: #ecf0f1;
            color: #2c3e50;
            padding: 3px 8px;
            border-radius: 12px;
            font-size: 0.75rem;
            font-weight: 500;
        }
        
        .intel-footer {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding-top: 15px;
            border-top: 1px solid #ecf0f1;
            font-size: 0.9rem;
            color: #7f8c8d;
        }
        
        .intel-date {
            font-style: italic;
        }
        
        .intel-link {
            color: #3498db;
            text-decoration: none;
            font-weight: 600;
        }
        
        .intel-link:hover {
            text-decoration: underline;
        }
        
        .error {
            background: #e74c3c;
            color: white;
            padding: 15px;
            border-radius: 8px;
            margin: 20px 0;
            text-align: center;
            font-weight: 600;
        }
        
        .empty-state {
            text-align: center;
            padding: 60px 20px;
            color: #7f8c8d;
        }
        
        .empty-state h3 {
            margin-bottom: 10px;
            font-size: 1.5rem;
        }
        
        @media (max-width: 768px) {
            .container {
                padding: 10px;
            }
            
            .header h1 {
                font-size: 2rem;
            }
            
            .intel-grid {
                grid-template-columns: 1fr;
            }
            
            .intel-header {
                flex-direction: column;
                align-items: flex-start;
            }
            
            .intel-meta {
                align-items: flex-start;
                flex-direction: row;
                flex-wrap: wrap;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <!-- Header -->
        <div class="header">
            <h1>🛡️ Cyber Threat Intelligence Aggregator</h1>
            <p>Comprehensive threat intelligence from multiple premium security sources</p>
            <div class="source-tags">
                <span class="source-tag">🏛️ Government Sources</span>
                <span class="source-tag">🏢 Vendor Advisories</span>
                <span class="source-tag">🔍 Threat Research</span>
                <span class="source-tag">🚨 IOC Feeds</span>
            </div>
        </div>
        
        <!-- Controls -->
        <div class="controls">
            <div style="text-align: center; margin-bottom: 20px;">
                <button class="btn btn-info" onclick="toggleSourcesView()">
                    📋 View All Sources
                </button>
            </div>
            
            <div id="sourcesView" class="sources-view">
                <h3 style="margin-bottom: 15px; color: #2c3e50;">📡 Configured Intelligence Sources</h3>
                <div id="sourcesList" class="sources-grid">
                    <!-- Sources will be loaded here -->
                </div>
            </div>
            
            <div class="button-group">
                <button class="btn btn-primary" onclick="collectIntelligence()">
                    🔄 Collect Intelligence
                </button>
                <button class="btn btn-success" onclick="exportData('excel')" disabled id="exportExcel">
                    📊 Export to Excel
                </button>
                <button class="btn btn-warning" onclick="exportData('json')" disabled id="exportJson">
                    📄 Export to JSON
                </button>
            </div>
            
            <div class="filters">
                <div class="filter-group">
                    <label for="categoryFilter">Category:</label>
                    <select id="categoryFilter" onchange="applyFilters()">
                        <option value="">All Categories</option>
                        <option value="threat_intel">Threat Intel</option>
                        <option value="advisory">Advisories</option>
                        <option value="ioc">IOCs</option>
                    </select>
                </div>
                
                <div class="filter-group">
                    <label for="sourceFilter">Source:</label>
                    <select id="sourceFilter" onchange="applyFilters()">
                        <option value="">All Sources</option>
                    </select>
                </div>
                
                <div class="filter-group">
                    <label for="severityFilter">Severity:</label>
                    <select id="severityFilter" onchange="applyFilters()">
                        <option value="">All Severities</option>
                        <option value="critical">Critical</option>
                        <option value="high">High</option>
                        <option value="medium">Medium</option>
                        <option value="low">Low</option>
                        <option value="unknown">Unknown</option>
                    </select>
                </div>
                
                <div class="filter-group">
                    <label for="searchFilter">Search:</label>
                    <input type="text" id="searchFilter" placeholder="Search..." oninput="applyFilters()">
                </div>
            </div>
        </div>
        
        <!-- Statistics -->
        <div class="stats" id="statsSection" style="display: none;">
            <div class="stats-grid" id="statsGrid">
                <!-- Stats will be populated here -->
            </div>
        </div>
        
        <!-- Results -->
        <div class="results">
            <h2>📋 Intelligence Results</h2>
            <div id="resultsContent">
                <div class="empty-state">
                    <h3>🎯 Ready to Collect Intelligence</h3>
                    <p>Click "Collect Intelligence" to start gathering threat data from multiple sources</p>
                </div>
            </div>
        </div>
    </div>

    <script>
        let currentData = [];
        let allSources = [];
        let sourceStats = {};
        
        // Load sources on page load
        document.addEventListener('DOMContentLoaded', function() {
            loadSources();
        });
        
        async function loadSources() {
            try {
                const response = await fetch('/api/sources');
                const data = await response.json();
                
                if (data.success) {
                    displaySources(data.sources);
                }
            } catch (error) {
                console.error('Error loading sources:', error);
            }
        }
        
        function displaySources(sources) {
            const sourcesList = document.getElementById('sourcesList');
            
            let html = '';
            sources.forEach(source => {
                const typeColor = {
                    'rss': '#3498db',
                    'nvd_api': '#e74c3c',
                    'threatfox_api': '#e74c3c',
                    'malware_bazaar_api': '#e74c3c'
                }[source.type] || '#95a5a6';
                
                html += `
                    <div class="source-item" style="border-left-color: ${typeColor};">
                        <div class="source-name">${source.description}</div>
                        <div class="source-meta">
                            <span style="background: ${typeColor}; color: white;">${source.type.replace('_', ' ').toUpperCase()}</span>
                            <span>${source.category.replace('_', ' ').toUpperCase()}</span>
                        </div>
                    </div>
                `;
            });
            
            sourcesList.innerHTML = html;
        }
        
        function toggleSourcesView() {
            const sourcesView = document.getElementById('sourcesView');
            sourcesView.style.display = sourcesView.style.display === 'none' ? 'block' : 'none';
        }
        
        async function collectIntelligence() {
            const resultsContent = document.getElementById('resultsContent');
            const statsSection = document.getElementById('statsSection');
            
            // Show loading state
            resultsContent.innerHTML = `
                <div class="loading">
                    <div class="spinner"></div>
                    <p>🌍 Collecting threat intelligence from multiple premium sources...</p>
                    <p><small>📡 CISA • Microsoft • CrowdStrike • Kaspersky • Palo Alto • ThreatFox • MalwareBazaar</small></p>
                    <p><small>⏱️ This comprehensive collection may take 1-2 minutes</small></p>
                </div>
            `;
            
            // Hide stats during loading
            statsSection.style.display = 'none';
            
            // Disable export buttons
            document.getElementById('exportExcel').disabled = true;
            document.getElementById('exportJson').disabled = true;
            
            try {
                const response = await fetch('/api/collect', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    }
                });
                
                const data = await response.json();
                
                if (data.success) {
                    currentData = data.items;
                    sourceStats = data.source_stats || {};
                    updateSourceFilter(Object.keys(data.sources));
                    displayResults(data);
                    displayStats(data);
                    
                    // Enable export buttons
                    document.getElementById('exportExcel').disabled = false;
                    document.getElementById('exportJson').disabled = false;
                    
                    // Show stats
                    statsSection.style.display = 'block';
                } else {
                    throw new Error(data.error || 'Collection failed');
                }
                
            } catch (error) {
                resultsContent.innerHTML = `
                    <div class="error">
                        ❌ Error collecting intelligence: ${error.message}
                    </div>
                `;
                console.error('Collection error:', error);
            }
        }
        
        function displayStats(data) {
            const statsGrid = document.getElementById('statsGrid');
            
            const successfulSources = Object.values(sourceStats).filter(s => s.status === 'success').length;
            const totalSources = Object.keys(sourceStats).length;
            
            statsGrid.innerHTML = `
                <div class="stat-card">
                    <h3>${data.total_items}</h3>
                    <p>Total Items</p>
                </div>
                <div class="stat-card">
                    <h3>${data.duplicates_removed}</h3>
                    <p>Duplicates</p>
                </div>
                <div class="stat-card">
                    <h3>${successfulSources}/${totalSources}</h3>
                    <p>Sources</p>
                </div>
                <div class="stat-card">
                    <h3>${data.categories.threat_intel || 0}</h3>
                    <p>Threat Intel</p>
                </div>
                <div class="stat-card">
                    <h3>${data.categories.advisory || 0}</h3>
                    <p>Advisories</p>
                </div>
                <div class="stat-card">
                    <h3>${data.categories.ioc || 0}</h3>
                    <p>IOCs</p>
                </div>
                <div class="stat-card">
                    <h3>${data.severities.critical || 0}</h3>
                    <p>Critical</p>
                </div>
                <div class="stat-card">
                    <h3>${data.severities.high || 0}</h3>
                    <p>High</p>
                </div>
            `;
        }
        
        function displayResults(data) {
            const resultsContent = document.getElementById('resultsContent');
            
            if (data.items.length === 0) {
                resultsContent.innerHTML = `
                    <div class="empty-state">
                        <h3>📭 No Intelligence Found</h3>
                        <p>No threat intelligence items were collected from the sources.</p>
                    </div>
                `;
                return;
            }
            
            // Sort items by severity and date
            const sortedItems = data.items.sort((a, b) => {
                const severityOrder = { 'critical': 0, 'high': 1, 'medium': 2, 'low': 3, 'unknown': 4 };
                const severityDiff = severityOrder[a.severity] - severityOrder[b.severity];
                if (severityDiff !== 0) return severityDiff;
                
                return new Date(b.published_date) - new Date(a.published_date);
            });
            
            const itemsHtml = sortedItems.map(item => createIntelItemHtml(item)).join('');
            
            resultsContent.innerHTML = `
                <div style="margin-bottom: 15px; text-align: center; color: #7f8c8d;">
                    📊 Displaying ${data.items.length} threat intelligence items
                </div>
                <div class="intel-grid">
                    ${itemsHtml}
                </div>
            `;
        }
        
        function createIntelItemHtml(item) {
            const publishedDate = new Date(item.published_date).toLocaleDateString();
            const iocInfo = item.ioc_type && item.ioc_value ? 
                `<p><strong>IOC:</strong> ${item.ioc_type.toUpperCase()}: <code style="background:#f8f9fa;padding:2px 4px;border-radius:3px;font-family:monospace;">${item.ioc_value}</code></p>` : '';
            
            const sourceDisplay = item.source.replace(/_/g, ' ').replace(/\b\w/g, l => l.toUpperCase());
            
            const categoryIcons = {
                'threat_intel': '🎯',
                'advisory': '⚠️',
                'ioc': '🔍'
            };
            
            const maxDescLength = 300;
            const truncatedDesc = item.description.length > maxDescLength ? 
                item.description.substring(0, maxDescLength) + '...' : item.description;
            
            return `
                <div class="intel-item category-${item.category}">
                    <div class="intel-header">
                        <div class="intel-title">${categoryIcons[item.category] || '📄'} ${item.title}</div>
                        <div class="intel-meta">
                            <span class="category-badge">${item.category.replace('_', ' ')}</span>
                            <span class="severity-badge ${item.severity}">${item.severity}</span>
                            <span class="source-badge">${sourceDisplay}</span>
                        </div>
                    </div>
                    
                    <div class="intel-description">
                        ${truncatedDesc}
                        ${iocInfo}
                    </div>
                    
                    ${item.tags && item.tags.length > 0 ? `<div class="intel-tags">
                        ${item.tags.slice(0, 8).map(tag => `<span class="tag">#${tag}</span>`).join('')}
                        ${item.tags.length > 8 ? `<span class="tag" style="background:#95a5a6;color:white;">+${item.tags.length - 8} more</span>` : ''}
                    </div>` : ''}
                    
                    <div class="intel-footer">
                        <span class="intel-date">📅 ${publishedDate}</span>
                        ${item.url ? `<a href="${item.url}" target="_blank" class="intel-link">🔗 View Source</a>` : ''}
                    </div>
                </div>
            `;
        }
        
        function updateSourceFilter(sources) {
            allSources = sources;
            const sourceFilter = document.getElementById('sourceFilter');
            
            // Clear existing options except "All Sources"
            sourceFilter.innerHTML = '<option value="">All Sources</option>';
            
            // Add source options with better formatting
            sources.forEach(source => {
                const option = document.createElement('option');
                option.value = source;
                option.textContent = source.replace(/_/g, ' ').replace(/\b\w/g, l => l.toUpperCase());
                sourceFilter.appendChild(option);
            });
        }
        
        async function applyFilters() {
            const category = document.getElementById('categoryFilter').value;
            const source = document.getElementById('sourceFilter').value;
            const severity = document.getElementById('severityFilter').value;
            const search = document.getElementById('searchFilter').value;
            
            if (currentData.length === 0) return;
            
            try {
                const params = new URLSearchParams();
                if (category) params.append('category', category);
                if (source) params.append('source', source);
                if (severity) params.append('severity', severity);
                if (search) params.append('search', search);
                
                const response = await fetch(`/api/filter?${params.toString()}`);
                const data = await response.json();
                
                if (data.success) {
                    const resultsContent = document.getElementById('resultsContent');
                    
                    if (data.items.length === 0) {
                        resultsContent.innerHTML = `
                            <div class="empty-state">
                                <h3>🔍 No Results Found</h3>
                                <p>No items match your filter criteria. Try adjusting your filters.</p>
                                <p><small>Total available items: ${currentData.length}</small></p>
                            </div>
                        `;
                    } else {
                        const itemsHtml = data.items.map(item => createIntelItemHtml(item)).join('');
                        resultsContent.innerHTML = `
                            <div style="margin-bottom: 15px; text-align: center; color: #7f8c8d;">
                                Showing ${data.items.length} of ${currentData.length} items
                            </div>
                            <div class="intel-grid">
                                ${itemsHtml}
                            </div>
                        `;
                    }
                }
            } catch (error) {
                console.error('Filter error:', error);
            }
        }
        
        async function exportData(format) {
            try {
                const response = await fetch(`/api/export/${format}`);
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `cti_intelligence_${new Date().toISOString().split('T')[0]}.${format === 'excel' ? 'xlsx' : 'json'}`;
                    document.body.appendChild(a);
                    a.click();
                    window.URL.revokeObjectURL(url);
                    document.body.removeChild(a);
                    
                    // Show success message
                    const exportBtn = document.getElementById(format === 'excel' ? 'exportExcel' : 'exportJson');
                    const originalText = exportBtn.textContent;
                    exportBtn.textContent = '✅ Downloaded!';
                    setTimeout(() => {
                        exportBtn.textContent = originalText;
                    }, 2000);
                } else {
                    const error = await response.json();
                    throw new Error(error.error || 'Export failed');
                }
            } catch (error) {
                alert(`Export error: ${error.message}`);
                console.error('Export error:', error);
            }
        }
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    print("=" * 80)
    print("🛡️  CYBER THREAT INTELLIGENCE AGGREGATOR v2.0")
    print("=" * 80)
    print("🚀 Starting enterprise-grade web server...")
    print("🌐 Access the interface at: http://localhost:5000")
    print("=" * 80)
    print("📡 INTELLIGENCE SOURCES CONFIGURED:")
    print("   🏛️  Government: CISA, US-CERT, NVD")
    print("   🏢 Vendors: Microsoft, Cisco")
    print("   🔬 Research: CrowdStrike, Kaspersky, Palo Alto, Check Point, Malwarebytes")
    print("   📰 News: BleepingComputer, Krebs, Dark Reading, SecurityWeek, Hacker News")
    print("   🎯 IOCs: ThreatFox, MalwareBazaar")
    print("   🧠 Frameworks: SANS ISC")
    print("=" * 80)
    print("📊 Features:")
    print("   • 15+ Premium threat intelligence sources")
    print("   • Real-time intelligent deduplication")
    print("   • Advanced categorization (Threat Intel, Advisories, IOCs)")
    print("   • Severity-based prioritization")
    print("   • Source performance monitoring")
    print("   • Export to Excel/JSON with professional formatting")
    print("   • Advanced filtering and search capabilities")
    print("=" * 80)
    print("📦 Dependencies:")
    if RETRY_AVAILABLE:
        print("   ✅ HTTP retry functionality available")
    else:
        print("   ⚠️  HTTP retry not available - basic connectivity only")
    if EXCEL_AVAILABLE:
        print("   ✅ Excel export functionality available")
    else:
        print("   ⚠️  Excel export disabled - install openpyxl for Excel support")
    print("=" * 80)
    print("⚠️  IMPORTANT NOTES:")
    print("   • This tool collects data from public sources only")
    print("   • Collection may take 1-2 minutes depending on network")
    print("   • Please ensure compliance with data usage policies")
    print("   • Network connectivity required for all sources")
    print("=" * 80)
    
    try:
        app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
    except Exception as e:
        print(f"❌ Error starting server: {e}")
        print("💡 Make sure port 5000 is available and try again.")
        print("💡 Try running with administrator/sudo privileges if needed.")
