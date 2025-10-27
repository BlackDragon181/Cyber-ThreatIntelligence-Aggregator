import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import re
from datetime import datetime

class VulnerabilityPrioritizationTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Vulnerability Risk Prioritization Tool")
        self.root.geometry("900x700")
        
        self.input_file = None
        self.output_file = None
        
        self.setup_ui()
        
    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        title_label = ttk.Label(main_frame, text="Vulnerability Risk Prioritization Tool", 
                                font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=10)
        
        # File Selection
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(file_frame, text="Input File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.input_label = ttk.Label(file_frame, text="No file selected", foreground="gray")
        self.input_label.grid(row=0, column=1, sticky=tk.W, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.select_input_file).grid(row=0, column=2, padx=5)
        
        # Configuration Frame
        config_frame = ttk.LabelFrame(main_frame, text="Vulnerability Assessment Criteria", padding="10")
        config_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # CVSS Score
        ttk.Label(config_frame, text="CVSS Score Threshold (Priority 1):").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.cvss_threshold = ttk.Entry(config_frame, width=10)
        self.cvss_threshold.insert(0, "8.0")
        self.cvss_threshold.grid(row=0, column=1, sticky=tk.W, padx=5)
        
        # Affected Product
        ttk.Label(config_frame, text="Internet Facing:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.internet_facing = tk.BooleanVar(value=True)
        ttk.Checkbutton(config_frame, variable=self.internet_facing).grid(row=1, column=1, sticky=tk.W, padx=5)
        
        # RCE
        ttk.Label(config_frame, text="Is RCE (Remote Code Execution):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.is_rce = tk.BooleanVar(value=True)
        ttk.Checkbutton(config_frame, variable=self.is_rce).grid(row=2, column=1, sticky=tk.W, padx=5)
        
        # Privileges Required
        ttk.Label(config_frame, text="Requires Privileges:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.requires_privileges = tk.BooleanVar(value=False)
        ttk.Checkbutton(config_frame, variable=self.requires_privileges).grid(row=3, column=1, sticky=tk.W, padx=5)
        
        # User Interaction
        ttk.Label(config_frame, text="Requires User Interaction:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.user_interaction = tk.BooleanVar(value=False)
        ttk.Checkbutton(config_frame, variable=self.user_interaction).grid(row=4, column=1, sticky=tk.W, padx=5)
        
        # Exploit Available
        ttk.Label(config_frame, text="Exploit Publicly Available:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.exploit_available = tk.BooleanVar(value=True)
        ttk.Checkbutton(config_frame, variable=self.exploit_available).grid(row=5, column=1, sticky=tk.W, padx=5)
        
        # Wormable
        ttk.Label(config_frame, text="Is Wormable:").grid(row=6, column=0, sticky=tk.W, pady=5)
        self.is_wormable = tk.BooleanVar(value=True)
        ttk.Checkbutton(config_frame, variable=self.is_wormable).grid(row=6, column=1, sticky=tk.W, padx=5)
        
        # Likelihood
        ttk.Label(config_frame, text="High Likelihood for Exploitation:").grid(row=7, column=0, sticky=tk.W, pady=5)
        self.high_likelihood = tk.BooleanVar(value=True)
        ttk.Checkbutton(config_frame, variable=self.high_likelihood).grid(row=7, column=1, sticky=tk.W, padx=5)
        
        # IoC
        ttk.Label(config_frame, text="Indicators of Compromise Detected:").grid(row=8, column=0, sticky=tk.W, pady=5)
        self.ioc_detected = tk.BooleanVar(value=False)
        ttk.Checkbutton(config_frame, variable=self.ioc_detected).grid(row=8, column=1, sticky=tk.W, padx=5)
        
        # Process Button
        process_btn = ttk.Button(main_frame, text="Process Vulnerabilities", 
                                command=self.process_vulnerabilities, 
                                style='Accent.TButton')
        process_btn.grid(row=3, column=0, columnspan=3, pady=20)
        
        # Progress
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # Status
        self.status_label = ttk.Label(main_frame, text="Ready", foreground="green")
        self.status_label.grid(row=5, column=0, columnspan=3, pady=5)
        
        # Info Frame
        info_frame = ttk.LabelFrame(main_frame, text="Scoring Information", padding="10")
        info_frame.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        info_text = """
Score Ranges:
• Emergency (110-120): Immediate action required
• Critical (85-110): Address within 24-48 hours
• High (65-85): Address within 1 week
• Medium/Low (Below 65): Schedule for remediation
        """
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT).grid(row=0, column=0, sticky=tk.W)
        
    def select_input_file(self):
        filename = filedialog.askopenfilename(
            title="Select Vulnerability Report",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.input_file = filename
            self.input_label.config(text=filename.split('/')[-1], foreground="black")
            
    def calculate_priority_score(self, row):
        """Calculate priority score based on the scoring matrix"""
        score = 0
        priority_details = []
        
        # Extract CVSS score
        cvss_score = 0
        if 'CVSS' in row:
            cvss_str = str(row['CVSS'])
            match = re.search(r'\d+\.?\d*', cvss_str)
            if match:
                cvss_score = float(match.group())
        
        # Priority 1 factors (10 points each)
        # CVSS Score above 8.0
        if cvss_score > float(self.cvss_threshold.get()):
            score += 10
            priority_details.append("CVSS > 8.0 [10]")
        
        # Internet Facing/High Exposure
        if self.internet_facing.get():
            score += 10
            priority_details.append("Internet Facing [10]")
        
        # Availability Impact - High
        if 'availability' in str(row).lower() or cvss_score >= 7.0:
            score += 10
            priority_details.append("High Availability Impact [10]")
        
        # Integrity Impact - High
        if 'integrity' in str(row).lower() or cvss_score >= 7.0:
            score += 10
            priority_details.append("High Integrity Impact [10]")
        
        # Confidentiality Impact - High
        if 'confidentiality' in str(row).lower() or cvss_score >= 7.0:
            score += 10
            priority_details.append("High Confidentiality Impact [10]")
        
        # Is RCE
        if self.is_rce.get():
            score += 10
            priority_details.append("RCE Vulnerability [10]")
        
        # Privileges Required for exploitation
        if self.requires_privileges.get():
            score += 10
            priority_details.append("No Privileges Required [10]")
        
        # User Interaction required for exploitation
        if not self.user_interaction.get():
            score += 10
            priority_details.append("No User Interaction [10]")
        
        # Publicly available exploit code
        if self.exploit_available.get():
            score += 10
            priority_details.append("Public Exploit Available [10]")
        
        # Is Wormable
        if self.is_wormable.get():
            score += 10
            priority_details.append("Wormable [10]")
        
        # Likelihood for exploitation
        if self.high_likelihood.get():
            score += 10
            priority_details.append("High Exploitation Likelihood [10]")
        
        # Allowed events for Indicators of Compromise
        if self.ioc_detected.get():
            score += 10
            priority_details.append("IoC Detected [10]")
        
        return score, priority_details
    
    def get_risk_rating(self, score):
        """Determine risk rating based on score"""
        if score >= 110:
            return "Emergency"
        elif score >= 85:
            return "Critical"
        elif score >= 65:
            return "High"
        else:
            return "Medium/Low"
    
    def get_priority_level(self, score):
        """Determine priority level"""
        if score >= 110:
            return "Priority 1"
        elif score >= 65:
            return "Priority 2"
        else:
            return "Priority 3"
    
    def process_vulnerabilities(self):
        if not self.input_file:
            messagebox.showerror("Error", "Please select an input file first!")
            return
        
        try:
            self.progress.start()
            self.status_label.config(text="Processing...", foreground="orange")
            self.root.update()
            
            # Read input file
            if self.input_file.endswith('.csv'):
                df = pd.read_csv(self.input_file)
            else:
                df = pd.read_excel(self.input_file)
            
            # Check for required columns
            required_cols = ['CVE', 'IP', 'Asset']
            col_mapping = {}
            
            for col in required_cols:
                found = False
                for df_col in df.columns:
                    if col.lower() in df_col.lower():
                        col_mapping[col] = df_col
                        found = True
                        break
                if not found:
                    messagebox.showwarning("Warning", 
                        f"Column '{col}' not found. Processing will continue with available data.")
            
            # Calculate scores
            scores = []
            ratings = []
            priorities = []
            details_list = []
            
            for idx, row in df.iterrows():
                score, details = self.calculate_priority_score(row)
                rating = self.get_risk_rating(score)
                priority = self.get_priority_level(score)
                
                scores.append(score)
                ratings.append(rating)
                priorities.append(priority)
                details_list.append("; ".join(details))
            
            # Add new columns
            df['Risk_Score'] = scores
            df['Risk_Rating'] = ratings
            df['Priority_Level'] = priorities
            df['Scoring_Details'] = details_list
            
            # Sort by score (highest first)
            df = df.sort_values('Risk_Score', ascending=False)
            
            # Save output
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"Vulnerability_Prioritization_{timestamp}.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if output_file:
                # Create Excel writer with formatting
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Prioritized Vulnerabilities', index=False)
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Prioritized Vulnerabilities']
                    
                    # Apply conditional formatting colors
                    from openpyxl.styles import PatternFill, Font
                    
                    emergency_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    critical_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    high_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    medium_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    
                    # Find Risk_Rating column
                    rating_col = None
                    for idx, col in enumerate(df.columns, 1):
                        if col == 'Risk_Rating':
                            rating_col = idx
                            break
                    
                    if rating_col:
                        for row_idx in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row_idx, column=rating_col)
                            if cell.value == "Emergency":
                                cell.fill = emergency_fill
                                cell.font = Font(bold=True, color="FFFFFF")
                            elif cell.value == "Critical":
                                cell.fill = critical_fill
                                cell.font = Font(bold=True)
                            elif cell.value == "High":
                                cell.fill = high_fill
                                cell.font = Font(bold=True)
                            elif cell.value == "Medium/Low":
                                cell.fill = medium_fill
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                self.progress.stop()
                self.status_label.config(text=f"Processing complete! Output saved to: {output_file}", 
                                       foreground="green")
                
                # Show summary
                summary = f"""
Processing Complete!

Total Vulnerabilities: {len(df)}
Emergency: {len(df[df['Risk_Rating'] == 'Emergency'])}
Critical: {len(df[df['Risk_Rating'] == 'Critical'])}
High: {len(df[df['Risk_Rating'] == 'High'])}
Medium/Low: {len(df[df['Risk_Rating'] == 'Medium/Low'])}

Output file: {output_file}
                """
                messagebox.showinfo("Success", summary)
            else:
                self.progress.stop()
                self.status_label.config(text="Save cancelled", foreground="orange")
                
        except Exception as e:
            self.progress.stop()
            self.status_label.config(text=f"Error: {str(e)}", foreground="red")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = VulnerabilityPrioritizationTool(root)
    root.mainloop()
